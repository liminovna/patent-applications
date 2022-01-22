from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
import os
import shutil


def create_ipc_dir(IPC):

    try:
        os.mkdir(IPC)
    except FileExistsError:
        pass
    finally:
        os.chdir(IPC)


def create_fam_dir(family_num):

    try:
        os.mkdir(family_num)
    except FileExistsError:
        pass
    finally:
        os.chdir(family_num)


def open_page():

    link = info[4].hyperlink.target
    driver.get(link)

    try:
        driver.maximize_window()

    except:

        pass

    time.sleep(6)


def ru_find_the_pub():
    """
    if the publication linked in the table cannot provide full text of the description,
    then find and go to a RU publication that can
    """

    RU_pub = driver.find_element_by_class_name('search').text

    if 'A' in RU_pub[-2:]:

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((
            By.XPATH,
            '/html/body/div/div/div[2]/div/div[3]/div/div/div[2]/div[3]/section/div[2]/section/section[1]/div[5]')))
        elms = driver.find_elements_by_tag_name('a')

        for el in elms:

            if re.search('RU.+C?[1-3]?', el.text) or re.search('EA.+B?[1-3]?', el.text):

                webdriver.ActionChains(driver).move_to_element(el).click(el).perform()
                RU_pub = driver.find_element_by_class_name('search').text

                break

    return RU_pub

def ru_get_text():

    desc = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
        By.XPATH, '/html/body/div/div/div[2]/div/div[3]/div/div/div[2]/div[2]/ul/li[2]')))
    desc.click()

    time.sleep(5)

    elms = driver.find_elements_by_tag_name('button')

    for el in elms:

        if el.text == 'RU':

            try:

                el.click()

            except Exception as e:

                print(e)

            finally:

                time.sleep(3)
                text = driver.find_element_by_class_name('text-block__content--3zTqgrLE').text
                return text


def ru_write_text(RU_pub, ru_text):

    with open(f'{RU_pub}.txt', 'w+', encoding='utf8') as Ru_file:
        Ru_file.write(ru_text + '\n')
        # Ru_file.seek(0)
    print('\tru text done')

def other_pubs():

    pat_fam = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
        By.XPATH, '/html/body/div/div/div[2]/div/div[3]/div/div/div[2]/div[2]/ul/li[8]')))
    pat_fam.click()

    time.sleep(5)

def zh_to_pub():

    # look for WO publication

    pubs = driver.find_elements_by_class_name('with-focus--3bDmNng8')

    WO_found = False
    CN_found = False

    for WOpub in pubs:

        if re.search('WO.+[A-C]?[1-3]?', WOpub.text) and len(WOpub.text) < 20:

            time.sleep(3)
            WOpub.click()

            WO_found = True
            break

    if WO_found is False:

        # if no WO publication was found, look for CN publication
        pubs = driver.find_elements_by_class_name('with-focus--3bDmNng8')

        for CNpub in pubs:

            if re.search('CN.+C?[1-3]?', CNpub.text) and len(CNpub.text) < 20:

                time.sleep(3)
                CNpub.click()

                CN_found = True
                break

    if WO_found is False and CN_found is False:
        raise Exception


def zh_pub_num():

    try:

        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CLASS_NAME, 'search')))
        ZH_pub = driver.find_element_by_class_name('search').text

    except:

        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CLASS_NAME, 'h4--19TXYMyq')))
        ZH_pub = driver.find_element_by_class_name('h4--19TXYMyq').text

    return ZH_pub


def zh_get_text():

    desc = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
        By.XPATH, '/html/body/div/div/div[2]/div/div[3]/div/div/div[2]/div[2]/ul/li[2]')))
    desc.click()

    time.sleep(5)

    elms = driver.find_elements_by_tag_name('button')

    for el in elms:

        if el.text == 'ZH':

            try:

                el.click()

            except Exception as e:

                print(e)

            finally:
                time.sleep(3)
                text = driver.find_element_by_class_name('text-block__content--3zTqgrLE').text

                return text


def zh_write_text(ZH_pub, zh_text):

    with open(f'{ZH_pub}.txt', 'w+', encoding='utf8') as Zh_file:
        Zh_file.write(zh_text + '\n')
        # Zh_file.seek(0)
        print('\tzh text done')


# initial data
PATH = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(PATH)
IPC = 'h04l-pt1'
xl_name = f'database/{IPC}.xlsx'
home_path = os.getcwd()

wb = load_workbook(xl_name)
ws = wb['Results']
create_ipc_dir(IPC)

for info in ws.iter_rows(min_row=42, max_row=61): # for h04l-pt1
# for info in ws.iter_rows(min_row=61, max_row=86): # for g06-pt2

    os.chdir(os.path.join(home_path, IPC))
    print('Working on ' + '[' + str(info[10].value) + ']')
    create_fam_dir(family_num=str(info[10].value))
    open_page()
    try:
        RU_pub = ru_find_the_pub()
        ru_text = ru_get_text()
        ru_write_text(RU_pub=RU_pub, ru_text=ru_text)
        try:
            other_pubs()
            zh_to_pub()
            ZH_pub = zh_pub_num()
            zh_text = zh_get_text()
            zh_write_text(ZH_pub=ZH_pub, zh_text=zh_text)

        except:
            print('no zh publication for ' + str(info[10].value))
    except:
        print('no ru publication for ' + str(info[10].value))

    print('_'*30)

    time.sleep(5)

wb.close()
driver.quit()

