"""
1. из проранжированного списка глаголов выбрать глагол
2. найти его в общем списке слов и взять его частоту, вычислить ipm для слова в любой части речи
3. найти ipm этого слова в базе данных CLD и заменить точку на запятую
6. вставить в другую колонку
"""

from openpyxl import load_workbook
import time

wb = load_workbook('rfp.xlsx')
v_sheet = wb['verbs']
t_sheet = wb['total']
cld_sheet = wb['CLD']
rewr_sheet = wb['rwr']


# gets the verb
def get_word(i):

    word_pos = i[3].value
    word = word_pos[:word_pos.find('_')]

    return word


# calculates ipm of the word (disregarding POS) in the total wordlist
def get_ipm_patent(word):

    # for i in range(2, 1500):
    for i in range(1500, t_sheet.max_row):

        if word == t_sheet[f'C{i}'].value:

            raw_freq = t_sheet[f'B{i}'].value
            ipm_pat = raw_freq/956074*1000000

            return ipm_pat

    return -1


# get ipm from CLD
def get_ipm_CLD(word):

    for i in range(2, 40000):

        if word == cld_sheet[f'A{i}'].value:

            ipm_cld = cld_sheet[f'C{i}'].value.replace('.', ',')

            return ipm_cld

    return -1


# for i in v_sheet.iter_rows(min_row=2, max_row=v_sheet.max_row):

    # start_time = time.time()
    #
    # word = get_word(i)
    # ipm_pat = get_ipm_patent(word)
    # ipm_cld = get_ipm_CLD(word)
    #
    # with open('ipms.txt', 'a+', encoding='utf8') as outfile:
    #
    #     print(word, ipm_pat, ipm_cld, sep='\t', end='\n', file=outfile)
    #
    # print(word, '-----' + str(time.time() - start_time) + ' seconds -----')
    pass


for n in range(1, rewr_sheet.max_row):

    if rewr_sheet[f'B{n}'].value == -1:

        word = rewr_sheet[f'A{n}'].value
        ipm_pat = get_ipm_patent(word)

        rewr_sheet[f'B{n}'] = ipm_pat

wb.save('rfp2.xlsx')







